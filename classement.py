# -*- coding: utf-8 -*-
"""
Created on Mon Jan 18 18:17:06 2021

@author: LENOVO
"""

import pandas as pd
import openpyxl as opx
import numpy as np
import matplotlib.pyplot as plt
data=pd.read_csv("classement.csv",sep=";",header=0,encoding="Utf-8")
data.drop(['publications','citations','patents'],axis=1,inplace=True)
#print(data)

classement=data.groupby(['year'])['world_rank'].count()
dt=dict(Nombrede_Classement_par_année=classement)
df=pd.DataFrame(dt)
print(df)



def create_excel():
    nameF=input('Donner le nom du fichier excel\n')
    wp=opx.Workbook()
    #wp.create_sheet(name,index=None)
    wp.save(nameF+'.xlsx')
    return nameF

def save_excel(df,nameF, sheetname):
    writer=pd.ExcelWriter(nameF+'.xlsx', engine='openpyxl')
    writer.book=opx.load_workbook(nameF+'.xlsx')
    df.to_excel(nameF+'.xlsx',sheet_name=sheetname,header=True)
    writer.save
    
def save_plot(df,nameF,sheetname):
    writer=opx.load_workbook(nameF+'.xlsx')
    worksheet=writer.create_sheet(sheetname)
    img=opx.drawing.image.Image(df+'.png')
    worksheet.add_image(img)
    writer.save(nameF+'.xlsx')
    
def get_top10(data):
    df['data'] = pd.to_datetime(data[['year', 'institution']], errors = 'coerce')
df.head(10)
    
    

'''df['date'] = pd.to_datetime(df[['year', 'month', 'day']], errors = 'coerce')
df.head()'''

def get_national_top10(data,namep):
    data1=data[data.country==namep]
    data2=data1.sort_values(by=['national_rank'],ascending=True).head(10)
    return data2

def score(data,year):
    data1=data[data.year == year]
    maxi=data1[data1.score == data1['score'].max()]
    mini=data1[data1.score ==data1['score'].min()]
    moy=data1.groupby(['country'])['score'].mean()
    #data2=pd.Series([maxi[['institution','score']].values,mini[['institution','score']].values,moy],index=['Uni_maxi_score','Uni_min_score','moy_score_pays'])
    d={}
    d['maxi']=maxi[['institution','score']]
    d['mini']=mini[['institution','score']]
    d['moyenne']=moy
    data3=pd.Series(d)
    p=pd.DataFrame(data3)
    return p

def total_universities(data,year):
    data1=data[data.year == year]
    data2=data1.country.value_counts()
    return data2

def max_universities(data):
    d=[]
    for i in range(2012,2016):
        data1=data[data.year == i]['country'].value_counts().index.tolist()[0]
        d.append([i,data1])
    pad=pd.DataFrame(d,columns=['Year','Country'])
    return pad
    #data1=data.groupby(['year'])['country'].value_counts() 
    #return data1
def create_bar_garaphic(data,year):
    data1=data[data.year == year]
    x=data1.country.value_counts().index
    y=data1.country.value_counts().values
    #data2=data1.groupby(['country'])['institution'].count()
    plt.bar(x,y,color='blue')
    plt.xlabel("Pays")
    plt.ylabel("Nombre d'université")
    plt.title("le nombre des universités par pays, pour l'année"+str(year))
    name=input('Donner le nom de votre image pour la sauvegarde')
    plt.savefig(name)
    return name 
def create_scatter_garaphic(data,year,pays):
    data1=data[data.year == year]
    data2=data1[data1.country == pays]
    data3=data2[['institution','score']].head(5)
    plt.scatter(data3['institution'], data3['score'], color='blue')
    plt.xlabel('Institution')
    plt.ylabel('Score')
    plt.title('Top 5 université pour '+pays+' en '+str(year))
    name=input('Donner le nom de votre image pour la sauvegarde')
    plt.savefig(name)
    return name

#namef=create_excel()
print(get_top10(data))

#maprint(df1)
#save_excel(df1,namef,"mory")
"""data=pd.read_csv("classement.csv",sep=";",header=0,encoding="Utf-8")
data.drop(['publications','citations','patents'],axis=1,inplace=True)
print(data)
classement=data.groupby(['year'])['world_rank'].count()
dt=dict(Nombrede_Classement_par_année=classement)
df=pd.DataFrame(dt)
print(df)
print(score(data,2012))"""

######## Menu principal ########
'''while True:
    print("********************************************")
    print("************** Menu Principal **************")
    print("********************************************")
    data=pd.read_csv("classement.csv",sep=";",header=0,encoding="Utf-8")
    data.drop(['publications','citations','patents'],axis=1,inplace=True)
    print(data)
    classement=data.groupby(['year'])['world_rank'].count()
    dt=dict(Nombrede_Classement_par_année=classement)
    df=pd.DataFrame(dt)
    print(df)
    print("1- Créez un fichier excel sur le disque dur pour stocker les résultats de vos jeux de données \n ")
    print("2- Affichez les 10 premières universités de chaque année  classés selon 'world_rank' \n ")
    print("3- Saisissez le nom d'un pays: Pour savoir les 10 premières universités classées selon 'national_rank'\n")
    print("4- Saisissez une année: Pour savoir l'université ayant le maximum de score, l'université ayant le minimal de score, et la moyenne des scores, groupés par pays \n")
    print("5- Saisissez une année: Pour affichez le nombre total de classements par pays \n")
    print("6- Affichez pour chaque année le pays ayant le maximum de classements.\n")
    print("7- Affichez le graphique par batton répresentant le nombre des universités par pays, pour une année donnée.\n")
    print("8- Affichez le graphique par scatter répresentant le score des top 5 universités pour un pays et une année donnée \n")
    print("9- Quitter")
    while True:
        try:
            choix=int(input('Votre choix? \n'))
            break
        except ValueError:
            print("Saisissez un choix parmi les choix proposées\n")
        
    if choix == 1:
        name=create_excel()
    elif choix == 2:
        choix2=get_top10(data)
        sheetname=input("Veuilez donner un nom à votre feuille de calcul\n")
        save_excel(choix2, name, sheetname)
    elif choix == 3:
        namep=input("Donner le nom d'un pays\n")
        choix3=get_national_top10(data, namep)
        sheetname=input("Veuilez donner un nom à votre feuille de calcul\n")
        save_excel(choix3, name, sheetname)
    elif choix == 4:
        year=int(input("Donner l'année comprises entre 2012 et 2015\n"))
        choix4=score(data, year)
        sheetname=input("Veuilez donner un nom à votre feuille de calcul\n")
        save_excel(choix4, name, sheetname)
    elif choix == 5:
        year=int(input("Donner l'année comprises entre 2012 et 2015\n"))
        choix5=total_universities(data, year)
        sheetname=input("Veuilez donner un nom à votre feuille de calcul\n")
        save_excel(choix5, name, sheetname)
    elif choix == 6:
        choix6=max_universities(data)
        sheetname=input("Veuilez donner un nom à votre feuille de calcul\n")
        save_excel(choix6, name, sheetname)
    elif choix == 7 :
        year=int(input("Donner l'année comprises entre 2012 et 2015\n"))
        choix7=create_bar_garaphic(data, year)
        sheetname=input("Veuilez donner un nom à votre feuille de calcul\n")
        save_plot(choix7, name, sheetname)
    elif choix == 8:
        year=int(input("Donner l'année comprises entre 2012 et 2015\n"))
        pays=input("Donner le nom d'un pays\n")
        choix8=create_scatter_garaphic(data, year, pays)
        sheetname=input("Veuilez donner un nom à votre feuille de calcul\n")
        save_plot(choix8, name, sheetname)  
    elif choix == 9:
        break'''
    
